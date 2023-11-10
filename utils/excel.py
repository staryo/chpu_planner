import openpyxl


def excel_to_dict(excel_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(excel_path)

    # Select the first sheet
    if "Default" in workbook.sheetnames:
        worksheet = workbook["Default"]
    else:
        worksheet = workbook.active

    # Get the header row as a list
    header = [cell.value for cell in worksheet[1]]

    # Initialize an empty list to store the dictionaries
    data = []

    # Loop through the rows and create a dictionary for each row
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for j in range(len(header)):
            row_data[header[j]] = row[j]
        data.append(row_data)

    return data


def dict_to_excel(data, path: str) -> None:
    """

    @rtype: object
    """
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the header row
    header = list(data[0].keys())
    for i in range(len(header)):
        worksheet.cell(row=1, column=i + 1, value=header[i])

    # Write the data rows
    for i in range(len(data)):
        row_data = list(data[i].values())
        for j in range(len(row_data)):
            worksheet.cell(row=i + 2, column=j + 1, value=row_data[j])

    # Save the workbook to an Excel file
    workbook.save(path)
    workbook.close()
